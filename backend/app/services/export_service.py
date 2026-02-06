import csv
import io
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..models import Project, ExtractedValue
from ..exceptions import NotFoundError


class ExportService:
    """Generates CSV and Excel exports of extracted field values."""
    
    def __init__(self, db: Session):
        self.db = db
    
    def export_csv(self, project_id: int) -> tuple[str, bytes]:
        data = self._prepare_data(project_id)
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(data["headers"])
        for row in data["rows"]:
            writer.writerow(row)
        
        filename = f"{data['project_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return filename, output.getvalue().encode('utf-8')
    
    def export_excel(self, project_id: int) -> tuple[str, bytes]:
        data = self._prepare_data(project_id)
        
        wb = Workbook()
        self._create_data_sheet(wb, data)
        self._create_summary_sheet(wb, data)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{data['project_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return filename, output.getvalue()
    
    def _prepare_data(self, project_id: int) -> dict:
        project = self.db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise NotFoundError("Project", project_id)
        
        if not project.template:
            raise NotFoundError("Template", project.template_id)
        
        documents = [pd.document for pd in project.project_documents]
        fields = project.template.fields
        
        # Build headers: Field name + columns for each document
        headers = ["Field"]
        for doc in documents:
            name = doc.original_filename[:30]
            headers.extend([f"{name} - Value", f"{name} - Citation", f"{name} - Confidence", f"{name} - Status"])
        
        rows = []
        stats = {"total": 0, "approved": 0, "pending": 0, "rejected": 0, "edited": 0, "confidence_sum": 0}
        
        for field in fields:
            row = [field.field_label]
            for doc in documents:
                ev = self.db.query(ExtractedValue).filter(
                    ExtractedValue.document_id == doc.id,
                    ExtractedValue.template_field_id == field.id,
                    ExtractedValue.project_id == project_id
                ).first()
                
                if ev:
                    row.extend([
                        ev.normalized_value or ev.raw_value or "",
                        ev.citation or "",
                        f"{ev.confidence:.0%}" if ev.confidence else "0%",
                        ev.status or "pending"
                    ])
                    stats["total"] += 1
                    stats[ev.status] = stats.get(ev.status, 0) + 1
                    stats["confidence_sum"] += ev.confidence or 0
                else:
                    row.extend(["", "", "", ""])
            rows.append(row)
        
        stats["avg_confidence"] = stats["confidence_sum"] / stats["total"] if stats["total"] > 0 else 0
        
        return {
            "project_name": project.name.replace(" ", "_"),
            "template_name": project.template.name,
            "document_count": len(documents),
            "headers": headers,
            "rows": rows,
            "stats": stats
        }
    
    def _create_data_sheet(self, wb: Workbook, data: dict) -> None:
        ws = wb.active
        ws.title = "Extracted Fields"
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="top", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        status_fills = {
            "approved": PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid"),
            "rejected": PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid"),
            "edited": PatternFill(start_color="fefcbf", end_color="fefcbf", fill_type="solid"),
            "pending": PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid"),
        }
        
        # Write headers
        for col, header in enumerate(data["headers"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        
        # Write data rows
        for row_idx, row in enumerate(data["rows"], 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
                cell.border = border
                
                # Color-code status columns
                header = data["headers"][col_idx - 1]
                if "Status" in header and value in status_fills:
                    cell.fill = status_fills[value]
        
        # Column widths
        for col in range(1, len(data["headers"]) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.column_dimensions['A'].width = 25
        ws.freeze_panes = 'B2'
    
    def _create_summary_sheet(self, wb: Workbook, data: dict) -> None:
        ws = wb.create_sheet("Summary")
        summary = [
            ["Project Name", data["project_name"]],
            ["Template", data["template_name"]],
            ["Document Count", data["document_count"]],
            ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Fields", data["stats"]["total"]],
            ["Approved", data["stats"]["approved"]],
            ["Pending", data["stats"]["pending"]],
            ["Rejected", data["stats"]["rejected"]],
            ["Edited", data["stats"]["edited"]],
            ["Average Confidence", f"{data['stats']['avg_confidence']:.1%}"],
        ]
        
        for row_idx, (label, value) in enumerate(summary, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
